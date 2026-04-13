import streamlit as st
from datetime import datetime, timedelta
import json
import os
import pandas as pd

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet

from streamlit_drawable_canvas import st_canvas
from PIL import Image as PILImage

from openpyxl import Workbook, load_workbook


# =========================
# 📱 MOBILE OPTIMISÉ
# =========================
st.set_page_config(
    page_title="V17 PRO MAX",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@media (max-width: 768px){
    .block-container{
        padding: 10px !important;
    }
}
</style>
""", unsafe_allow_html=True)


st.title("📋 CHANTIER V17 PRO MAX")


# =========================
# 📂 HISTORIQUE
# =========================
file_json = "historique.json"

if os.path.exists(file_json):
    with open(file_json, "r") as f:
        historique = json.load(f)
else:
    historique = []


# =========================
# 📅 DATE FR
# =========================
st.subheader("📅 Date du chantier")

date_obj = st.date_input("Choisir la date", value=datetime.now())

jours = ["lundi","mardi","mercredi","jeudi","vendredi","samedi","dimanche"]
mois = ["janvier","février","mars","avril","mai","juin","juillet","août",
        "septembre","octobre","novembre","décembre"]

date_jour = f"{jours[date_obj.weekday()]} {date_obj.day} {mois[date_obj.month-1]} {date_obj.year}"

st.write("📅", date_jour)


# =========================
# 🏗 DONNÉES
# =========================
chantier = st.text_input("🏗 Chantier")
localisation = st.text_input("📍 Localisation")
engin = st.text_input("🚜 Engin")
travail = st.text_area("🧾 Travail")


# =========================
# ⏱ HORAIRES
# =========================
col1, col2 = st.columns(2)

with col1:
    debut_matin = st.time_input("Début matin")
    fin_matin = st.time_input("Fin matin")

with col2:
    debut_aprem = st.time_input("Début aprem")
    fin_aprem = st.time_input("Fin aprem")


def calc(d1, d2):
    if d1 and d2:
        a = datetime.combine(datetime.today(), d1)
        b = datetime.combine(datetime.today(), d2)
        if b > a:
            return b - a
    return timedelta(0)


def fmt(td):
    return f"{td.seconds//3600}h{(td.seconds%3600)//60:02d}"


total = calc(debut_matin, fin_matin) + calc(debut_aprem, fin_aprem)

st.write("🕒 Total :", fmt(total))


# =========================
# ✍️ SIGNATURE
# =========================
st.subheader("✍️ Signature")

canvas = st_canvas(
    stroke_width=3,
    stroke_color="black",
    background_color="white",
    height=150,
    width=400,
    drawing_mode="freedraw",
    key="sig"
)

signature_path = None

if canvas.image_data is not None:
    img = PILImage.fromarray(canvas.image_data.astype("uint8"))
    signature_path = "signature.png"
    img.save(signature_path)


# =========================
# 📸 PHOTOS
# =========================
st.subheader("📸 Photos")

photos = st.file_uploader(
    "Ajouter photos",
    type=["jpg","jpeg","png"],
    accept_multiple_files=True
)

legendes = []

if photos:
    for i, p in enumerate(photos):
        leg = st.text_input(f"Légende {i+1}", key=f"leg{i}")
        legendes.append(leg)


# =========================
# 📦 DATA
# =========================
data = {
    "date": date_jour,
    "chantier": chantier,
    "localisation": localisation,
    "engin": engin,
    "travail": travail,
    "heures": fmt(total),
    "timestamp": datetime.now().isoformat()
}


# =========================
# 💾 SAUVEGARDE
# =========================
if st.button("💾 ENREGISTRER"):
    historique.append(data)

    with open(file_json, "w") as f:
        json.dump(historique, f, indent=4)

    st.success("Sauvegarde OK ✔")
    st.rerun()


# =========================
# 📊 STATISTIQUES
# =========================
st.subheader("📊 Statistiques")

if historique:

    df = pd.DataFrame(historique)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")

    def to_decimal(h):
        try:
            a,b = h.split("h")
            return float(a)+float(b)/60
        except:
            return 0

    df["heures_num"] = df["heures"].apply(to_decimal)

    df["semaine"] = df["date"].dt.isocalendar().week
    st.write("📅 Heures/semaine")
    st.bar_chart(df.groupby("semaine")["heures_num"].sum())

    df["mois"] = df["date"].dt.to_period("M").astype(str)
    st.write("📆 Heures/mois")
    st.bar_chart(df.groupby("mois")["heures_num"].sum())

else:
    st.info("Aucune donnée")


# =========================
# 📄 PDF
# =========================
def generate_pdf():

    doc = SimpleDocTemplate("rapport.pdf")
    styles = getSampleStyleSheet()
    e = []

    e.append(Paragraph("RAPPORT CHANTIER V17 PRO MAX", styles["Title"]))
    e.append(Spacer(1, 20))

    e.append(Paragraph(f"Date : {data['date']}", styles["Normal"]))
    e.append(Paragraph(f"Chantier : {data['chantier']}", styles["Normal"]))
    e.append(Paragraph(f"Localisation : {data['localisation']}", styles["Normal"]))
    e.append(Paragraph(f"Engin : {data['engin']}", styles["Normal"]))
    e.append(Paragraph(f"Heures : {data['heures']}", styles["Normal"]))

    e.append(Spacer(1, 10))
    e.append(Paragraph("Travail :", styles["Heading2"]))
    e.append(Paragraph(data["travail"], styles["Normal"]))

    if signature_path:
        e.append(Spacer(1, 20))
        e.append(Image(signature_path, width=200, height=100))

    if photos:
        e.append(Spacer(1, 20))
        e.append(Paragraph("Photos :", styles["Heading2"]))

        for i,p in enumerate(photos):
            img = PILImage.open(p)
            img.thumbnail((800,800))
            path = f"p{i}.jpg"
            img.save(path)
            e.append(Image(path, width=300, height=200))

            if i < len(legendes):
                e.append(Paragraph(legendes[i], styles["Normal"]))

    doc.build(e)


if st.button("📄 PDF"):
    generate_pdf()
    with open("rapport.pdf","rb") as f:
        st.download_button("Télécharger PDF", f, "rapport.pdf")


# =========================
# 📊 EXCEL
# =========================
def to_decimal(h):
    try:
        a,b=h.split("h")
        return float(a)+float(b)/60
    except:
        return 0


if st.button("📊 EXCEL"):

    file="chantier.xlsx"

    if not os.path.exists(file):
        wb=Workbook()
        ws=wb.active
        ws.append(["Date","Chantier","Localisation","Engin","Travail","Heures"])
        wb.save(file)

    wb=load_workbook(file)
    ws=wb.active

    ws.append([
        data["date"],
        data["chantier"],
        data["localisation"],
        data["engin"],
        data["travail"],
        to_decimal(data["heures"])
    ])

    wb.save(file)

    with open(file,"rb") as f:
        st.download_button("Télécharger Excel", f, "chantier.xlsx")


# =========================
# 📂 HISTORIQUE + SUPPRESSION
# =========================
st.subheader("📂 Historique")

st.write(historique)

if st.button("🧹 Effacer historique"):
    if os.path.exists(file_json):
        os.remove(file_json)
    st.success("Supprimé ✔")
    st.rerun()